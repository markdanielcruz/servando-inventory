import streamlit as st
import gspread
from google.oauth2.service_account import Credentials
from datetime import datetime, date
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Servando Warehouse Inventory",
    page_icon="🏪",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── CSS ────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Cormorant+Garamond:wght@400;600;700&family=Inter:wght@300;400;500;600&display=swap');

*, *::before, *::after { box-sizing: border-box; }
html, body, [class*="css"] {
    font-family: 'Inter', sans-serif;
    background-color: #0A1208;
    color: #D4D0C8;
}
.stApp {
    background-color: #0A1208;
    background-attachment: fixed;
    background-size: cover;
}
[data-testid="stSidebar"] {
    background-color: rgba(8,14,8,0.95) !important;
    border-right: 1px solid #2A3828;
    backdrop-filter: blur(8px);
}
[data-testid="stSidebar"] * { color: #D4D0C8 !important; }
.main-header {
    background: linear-gradient(135deg, rgba(26,46,26,0.85) 0%, rgba(21,37,21,0.85) 100%);
    border: 1px solid #2A3828;
    border-radius: 12px;
    padding: 20px 28px;
    margin-bottom: 20px;
    backdrop-filter: blur(6px);
}
.main-header h1 {
    font-family: 'Cormorant Garamond', serif;
    font-size: 1.7rem;
    color: #8CAF7A;
    margin: 0 0 2px 0;
    font-weight: 600;
}
.main-header p { font-size: 0.7rem; letter-spacing: 2.5px; text-transform: uppercase; color: #4A6B3E; margin: 0; }
.card {
    background: #111E11;
    border: 1px solid #1E2E1C;
    border-radius: 12px;
    padding: 1.25rem 1.5rem;
    margin-bottom: 1rem;
}
.section-title {
    font-size: 0.62rem;
    font-weight: 600;
    letter-spacing: 3px;
    text-transform: uppercase;
    color: #5A7A52;
    margin-bottom: 0.75rem;
    padding-bottom: 6px;
    border-bottom: 1px solid #1E2E1C;
    display: block;
}
.stButton > button {
    background-color: #2A3E28 !important;
    color: #A8C896 !important;
    border: 1px solid #3A5238 !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    font-size: 0.82rem !important;
    transition: all 0.2s !important;
}
.stButton > button:hover {
    background-color: #3A5238 !important;
    border-color: #5A7A52 !important;
    color: #C8DCC0 !important;
}
.stTextInput > div > div > input,
.stNumberInput > div > div > input,
.stTextArea > div > textarea,
.stSelectbox > div > div {
    background-color: #111E11 !important;
    border: 1px solid #2A3828 !important;
    border-radius: 8px !important;
    color: #D4D0C8 !important;
    font-size: 0.88rem !important;
}
label { color: #8A9E84 !important; font-size: 0.78rem !important; font-weight: 500 !important; }
.log-entry {
    background: #0E1810;
    border-left: 3px solid #3A5238;
    padding: 0.6rem 1rem;
    border-radius: 0 8px 8px 0;
    margin-bottom: 0.4rem;
    font-size: 0.88rem;
    color: #D4D0C8;
}
.po-item-row {
    background: #0E1810;
    border: 1px solid #2A3828;
    border-radius: 8px;
    padding: 0.6rem 1rem;
    margin-bottom: 0.4rem;
    color: #D4D0C8;
}
.info-box {
    background: #111E11;
    border: 1px solid #2A3828;
    border-radius: 8px;
    padding: 0.6rem 1rem;
    color: #8CAF7A;
    margin: 0.4rem 0;
    font-size: 0.88rem;
}
.success-box {
    background: #1A2E1A;
    border: 1px solid #3A5238;
    border-radius: 8px;
    padding: 0.6rem 1rem;
    color: #A8C896;
    margin: 0.4rem 0;
}
[data-testid="metric-container"] {
    background: #111E11 !important;
    border: 1px solid #1E2E1C !important;
    border-radius: 10px !important;
    padding: 12px 14px !important;
}
[data-testid="metric-container"] label {
    font-size: 0.6rem !important; letter-spacing: 2px !important;
    text-transform: uppercase !important; color: #5A7A52 !important;
}
[data-testid="stMetricValue"] {
    font-family: 'Cormorant Garamond', serif !important;
    font-size: 1.5rem !important; font-weight: 600 !important; color: #A8C896 !important;
}
[data-testid="stDataFrame"] { border: 1px solid #1E2E1C !important; border-radius: 10px !important; }
hr { border-color: #1E2E1C !important; }
h3 { font-family: 'Cormorant Garamond', serif !important; color: #A8C896 !important; }
</style>
""", unsafe_allow_html=True)

# ── Constants ──────────────────────────────────────────────────────────────────
SPREADSHEET_ID = "1pYqecDqe_qUwkexRiG5mmOaYV58_1uoRCx_Yy4CepNQ"
SCOPES = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]

ITEMS_SHEET    = "v2_ITEMS"
LOG_SHEET      = "v2_LOG"

ITEMS_HEADERS  = ["ITEM", "UNIT COST", "UNIT OF MEASURE", "CATEGORY", "BEGINNING_STOCKS", "ACTIVE"]
# NOTE ON MIGRATION: RESTAURANT and BANQUET were renamed to RESTO and KITCHEN_BANQUET
# (same column position — only the header text changes). KITCHEN_ALACARTE and
# DAMAGED_OUT are brand-new columns appended at the very end so no existing data
# shifts position. Before deploying this, open the v2_LOG sheet and:
#   1. Rename cell J1 (header for old "BANQUET") to KITCHEN_BANQUET
#   2. Rename cell J1... i.e. the column that says RESTAURANT to RESTO
#   3. Add two new header cells at the end (after NOTES): KITCHEN_ALACARTE, DAMAGED_OUT
LOG_HEADERS    = ["TIMESTAMP", "DATE", "MONTH", "ITEM", "STAFF", "TXN_TYPE", "REF_NUMBER",
                  "ADD_IN", "OVER", "RESTO", "KITCHEN_BANQUET", "CAFE", "BAR", "OTHERS", "SPOILAGE", "NOTES",
                  "KITCHEN_ALACARTE", "DAMAGED_OUT"]

DEPARTMENTS    = ["Resto", "Café", "Bar", "Kitchen Ala Carte", "Kitchen Banquet", "Others"]
DEPT_COL_MAP   = {"Resto": "RESTO", "Kitchen Banquet": "KITCHEN_BANQUET", "Café": "CAFE",
                  "Bar": "BAR", "Kitchen Ala Carte": "KITCHEN_ALACARTE", "Others": "OTHERS"}
DEPT_LOG_INDEX = {"Resto": 9, "Kitchen Banquet": 10, "Café": 11, "Bar": 12, "Kitchen Ala Carte": 16}
DEPT_DISPLAY   = {"RESTO": "Resto", "KITCHEN_BANQUET": "Kitchen Banquet", "CAFE": "Café",
                  "BAR": "Bar", "KITCHEN_ALACARTE": "Kitchen Ala Carte", "OTHERS": "Others"}
CATEGORIES     = ["beverage","beef","chicken","pork","seafood","fresh","dry","wet","rtc","meat","frozen","dessert","other"]

ADJ_TYPES      = ["Over (Add +)", "Spoilage (Remove −)", "Damaged (Remove −)"]

LOGO_B64 = "PASTE_YOUR_ORIGINAL_LOGO_B64_HERE"  # unchanged by this update; copy your real value back in from your current app.py

# ── Google Sheets ──────────────────────────────────────────────────────────────
@st.cache_resource
def get_client():
    creds = Credentials.from_service_account_info(dict(st.secrets["gcp_service_account"]), scopes=SCOPES)
    return gspread.authorize(creds)

def get_spreadsheet():
    return get_client().open_by_key(SPREADSHEET_ID)

def ensure_sheet(ss, name, headers):
    try:
        ws = ss.worksheet(name)
    except gspread.WorksheetNotFound:
        ws = ss.add_worksheet(title=name, rows=5000, cols=len(headers)+2)
        ws.append_row(headers)
    return ws

@st.cache_data(ttl=30)
def load_items(_ss_id):
    ss = get_spreadsheet()
    ws = ensure_sheet(ss, ITEMS_SHEET, ITEMS_HEADERS)
    data = ws.get_all_records()
    return pd.DataFrame(data) if data else pd.DataFrame(columns=ITEMS_HEADERS)

@st.cache_data(ttl=30)
def load_log(_ss_id):
    ss = get_spreadsheet()
    ws = ensure_sheet(ss, LOG_SHEET, LOG_HEADERS)
    data = ws.get_all_records()
    return pd.DataFrame(data) if data else pd.DataFrame(columns=LOG_HEADERS)

def invalidate_cache():
    load_items.clear()
    load_log.clear()

def num(val):
    try: return float(val or 0)
    except: return 0.0

def replace_ref_rows(ss, ref, new_rows):
    """
    Clean-replace: remove every log row belonging to `ref`, then append `new_rows`
    (list of lists matching LOG_HEADERS order) in their place. Pass new_rows=[]
    to delete the reference entirely.
    """
    ws = ensure_sheet(ss, LOG_SHEET, LOG_HEADERS)
    all_records = ws.get_all_records()
    kept = [r for r in all_records if str(r.get("REF_NUMBER", "")) != ref]
    ws.clear()
    ws.append_row(LOG_HEADERS)
    if kept:
        ws.append_rows([[r.get(h, "") for h in LOG_HEADERS] for r in kept])
    if new_rows:
        ws.append_rows(new_rows)
    invalidate_cache()


def build_doc_xlsx(doc_type, ref, date_str, staff_str, meta3_label, meta3_val, rows_data, logo_b64_str):
    """
    Build a PO or Delivery xlsx matching the exact template format.
    rows_data: list of (num, item_name, unit, qty, amount)
    Returns BytesIO buffer.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    from io import BytesIO
    import base64

    wb = Workbook()
    ws = wb.active
    ws.title = doc_type

    # ── fonts ──
    f14b  = Font(bold=True, name="Arial", size=14)
    f8g   = Font(name="Arial", size=8,  color="888888")
    f12g  = Font(bold=True, name="Arial", size=12, color="888888")
    f12b  = Font(bold=True, name="Arial", size=12)
    f9w   = Font(bold=True, name="Arial", size=9,  color="FFFFFF")
    f11   = Font(name="Arial", size=11)
    f11b  = Font(bold=True, name="Arial", size=11)
    hdr_fill = PatternFill("solid", fgColor="1A2E1A")
    tot_fill = PatternFill("solid", fgColor="E8F0E8")
    ctr  = Alignment(horizontal="center")
    rgt  = Alignment(horizontal="right")
    vctr = Alignment(vertical="center")

    # ── col widths / row heights ──
    ws.column_dimensions["A"].width = 5.0
    ws.column_dimensions["B"].width = 38.0
    ws.column_dimensions["C"].width = 10.90625
    ws.column_dimensions["D"].width = 12.0
    ws.column_dimensions["E"].width = 16.0

    ws.row_dimensions[1].height = 22.0
    ws.row_dimensions[2].height = 14.0
    ws.row_dimensions[3].height = 11.0
    ws.row_dimensions[4].height = 18.0
    ws.row_dimensions[5].height = 16.0
    ws.row_dimensions[6].height = 8.0
    ws.row_dimensions[7].height = 18.0

    # ── logo ──
    try:
        logo_buf = BytesIO(base64.b64decode(logo_b64_str))
        xl_logo = XLImage(logo_buf)
        xl_logo.width = 220
        xl_logo.height = 49
        xl_logo.anchor = "D1"
        ws.add_image(xl_logo)
    except Exception:
        pass

    # ── row 1: title ──
    ws.merge_cells("A1:E1")
    ws["A1"] = doc_type
    ws["A1"].font = f14b
    ws["A1"].alignment = vctr

    # ── row 2: reference ──
    ws.merge_cells("A2:E2")
    ws["A2"] = f"Reference: {ref}"
    ws["A2"].font = f8g

    # ── rows 4–5: meta labels + values ──
    meta1_label = "Prepared By" if doc_type == "PURCHASE ORDER" else "Received By"
    ws["B4"] = meta1_label;  ws["B4"].font = f12g
    ws["C4"] = "Date";       ws["C4"].font = f12g
    ws["E4"] = meta3_label;  ws["E4"].font = f12g
    ws["B5"] = staff_str;    ws["B5"].font = f12b
    ws["C5"] = date_str;     ws["C5"].font = f12b
    ws["E5"] = meta3_val;    ws["E5"].font = f12b

    # ── row 7: table header ──
    for ci, (hdr, al) in enumerate([
        ("#", ctr), ("Item", None), ("Unit", ctr), ("Qty", ctr), ("Amount", ctr)
    ], 1):
        c = ws.cell(row=7, column=ci, value=hdr)
        c.font = f9w
        c.fill = hdr_fill
        c.alignment = al or Alignment()

    # ── rows 8–40: data rows (30 slots) ──
    for ri in range(30):
        row_num = 8 + ri
        ws.row_dimensions[row_num].height = 18.0 if ri == 0 else None
        if ri < len(rows_data):
            num_i, item_n, unit, qty, amt = rows_data[ri]
            qty_val = int(qty) if qty == int(qty) else round(qty, 2)
            vals = [num_i, item_n, unit, qty_val, amt]
            aligns = [ctr, None, ctr, ctr, rgt]
            fmts   = [None, None, None, "#,##0", "₱#,##0.00"]
            for ci, (v, al, fmt) in enumerate(zip(vals, aligns, fmts), 1):
                c = ws.cell(row=row_num, column=ci, value=v)
                c.font = f11
                if al: c.alignment = al
                if fmt: c.number_format = fmt

    # ── row 41: total (always fixed row) ──
    ws["A41"].font = f11b
    ws["B41"].font = f11b
    ws["C41"].font = f11b
    ws["D41"].font = f11b
    ws.merge_cells("A41:D41")
    ws["A41"] = "TOTAL"
    ws["A41"].alignment = rgt
    total = sum(r[4] for r in rows_data)
    ws["E41"] = total
    ws["E41"].font = f11b
    ws["E41"].fill = tot_fill
    ws["E41"].alignment = rgt
    ws["E41"].number_format = "₱#,##0.00"

    # ── row 43: signature labels ──
    if doc_type == "PURCHASE ORDER":
        sigs = [("A43","Prepared By"), ("C43","Approved By"), ("E43","Received By")]
    else:
        sigs = [("A43","Received By"), ("C43","Checked By"),  ("E43","Noted By")]
    for coord, label in sigs:
        ws[coord] = label
        ws[coord].font = f8g

    # ── print setup ──
    ws.print_area = "A1:E43"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── Sidebar ────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style='background:linear-gradient(175deg,#1A2E1A 0%,#0E1C0E 60%,#080F08 100%);
                padding:24px 20px 18px;border-bottom:1px solid #2A3828;text-align:center;'>
        <div style='font-family:Cormorant Garamond,serif;font-size:1.5rem;font-weight:700;color:#8CAF7A;letter-spacing:2px;'>SERVANDO</div>
        <div style='font-size:0.56rem;letter-spacing:3px;text-transform:uppercase;color:#4A6B3E;margin-top:5px;'>Main Warehouse Inventory</div>
    </div>
    <div style='font-size:0.56rem;font-weight:600;letter-spacing:3px;text-transform:uppercase;
                color:#4A6B3E;padding:14px 12px 5px;border-bottom:1px solid #1E2E1C;margin-bottom:6px;'>
        Navigation
    </div>
    """, unsafe_allow_html=True)

    page = st.radio("", [
        "📊 Dashboard",
        "📥 Incoming Deliveries",
        "📋 Purchase Orders",
        "🔧 Stock Adjustment",
        "🔍 Item History",
        "📦 Items Master",
        "⬇️ Export to Excel",
        "⚙️ Setup",
    ], label_visibility="collapsed")

    st.markdown("---")
    if st.button("🔄 Refresh Data Now", use_container_width=True):
        invalidate_cache()
        st.success("Data refreshed!")
        st.rerun()

    st.markdown("---")
    st.markdown(f"<div style='font-size:0.62rem;letter-spacing:1.5px;text-transform:uppercase;color:#3A5238;'>{date.today().strftime('%B %d, %Y')}</div>", unsafe_allow_html=True)

# ── Connect ────────────────────────────────────────────────────────────────────
try:
    ss = get_spreadsheet()
except Exception as e:
    st.error(f"❌ Cannot connect to Google Sheets: {e}")
    st.stop()

# ══════════════════════════════════════════════════════════════════════════════
# PAGE: DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
if page == "📊 Dashboard":
    st.markdown('<div class="main-header"><h1>📊 Dashboard</h1><p>Live stock overview</p></div>', unsafe_allow_html=True)

    items_df = load_items(SPREADSHEET_ID)
    log_df   = load_log(SPREADSHEET_ID)

    if items_df.empty:
        st.info("No items yet. Go to ⚙️ Setup to initialize.")
        st.stop()

    active = items_df[items_df["ACTIVE"] == "YES"].copy()

    def compute_stock(item_name, beginning):
        if log_df.empty:
            return beginning
        ilog = log_df[log_df["ITEM"] == item_name]
        if ilog.empty:
            return beginning
        add_in   = ilog["ADD_IN"].apply(num).sum()
        over     = ilog["OVER"].apply(num).sum()
        resto    = ilog["RESTO"].apply(num).sum() if "RESTO" in ilog.columns else 0
        kbanq    = ilog["KITCHEN_BANQUET"].apply(num).sum() if "KITCHEN_BANQUET" in ilog.columns else 0
        kalac    = ilog["KITCHEN_ALACARTE"].apply(num).sum() if "KITCHEN_ALACARTE" in ilog.columns else 0
        cafe     = ilog["CAFE"].apply(num).sum()
        bar      = ilog["BAR"].apply(num).sum()
        others   = ilog["OTHERS"].apply(num).sum()
        spoil    = ilog["SPOILAGE"].apply(num).sum()
        damage   = ilog["DAMAGED_OUT"].apply(num).sum() if "DAMAGED_OUT" in ilog.columns else 0
        return beginning + add_in + over - resto - kbanq - kalac - cafe - bar - others - spoil - damage

    active["CURRENT_STOCK"] = active.apply(lambda r: compute_stock(r["ITEM"], num(r["BEGINNING_STOCKS"])), axis=1)
    active["TOTAL_WORTH"]   = active["CURRENT_STOCK"] * active["UNIT COST"].apply(num)

    total_worth  = active["TOTAL_WORTH"].sum()
    total_items  = len(active)
    low_stock    = len(active[active["CURRENT_STOCK"] <= 0])
    total_txns   = len(log_df) if not log_df.empty else 0

    c1, c2, c3, c4 = st.columns(4)
    with c1: st.metric("Total Items", total_items)
    with c2: st.metric("Total Stock Worth", f"₱{total_worth:,.2f}")
    with c3: st.metric("Zero / Negative Stock", low_stock)
    with c4: st.metric("Total Transactions", total_txns)

    st.markdown("---")

    zero = active[active["CURRENT_STOCK"] <= 0]
    if not zero.empty:
        st.markdown('<div class="section-title" style="color:#CC6A6A;">⚠️ Zero or Negative Stock Items</div>', unsafe_allow_html=True)
        st.dataframe(zero[["ITEM","UNIT OF MEASURE","CATEGORY","CURRENT_STOCK"]].sort_values("ITEM"),
                     use_container_width=True, hide_index=True)

# ══════════════════════════════════════════════════════════════════════════════
# PAGE: DELIVERIES
# ══════════════════════════════════════════════════════════════════════════════
elif page == "📥 Incoming Deliveries":
    st.markdown('<div class="main-header"><h1>📥 Incoming Deliveries</h1><p>Log incoming stock from suppliers</p></div>', unsafe_allow_html=True)

    items_df = load_items(SPREADSHEET_ID)
    active_items = [f'{r["ITEM"]} ({r["UNIT OF MEASURE"]})' for _, r in items_df[items_df["ACTIVE"]=="YES"].iterrows()] if not items_df.empty else []
    item_name_map = {f'{r["ITEM"]} ({r["UNIT OF MEASURE"]})': r["ITEM"] for _, r in items_df[items_df["ACTIVE"]=="YES"].iterrows()} if not items_df.empty else {}

    if not active_items:
        st.warning("No items found. Go to Setup first.")
        st.stop()

    if "delivery_cart" not in st.session_state:
        st.session_state.delivery_cart = []

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Delivery Details</div>', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1: delivery_date = st.date_input("📅 Delivery Date", value=date.today(), key="del_date")
    with c2: staff_name = st.text_input("👤 Received By", placeholder="Your name", key="del_staff")
    st.markdown('</div>', unsafe_allow_html=True)

    active_item_names_del = items_df[items_df["ACTIVE"] == "YES"]["ITEM"].tolist()
    item_unit_map_del = {r["ITEM"]: r["UNIT OF MEASURE"] for _, r in items_df[items_df["ACTIVE"]=="YES"].iterrows()}
    item_cost_map_del = {r["ITEM"]: num(r["UNIT COST"]) for _, r in items_df[items_df["ACTIVE"]=="YES"].iterrows()}

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Add Items to Delivery</div>', unsafe_allow_html=True)
    st.caption("Pick an item and quantity per row. Use the **+** at the bottom of the table to add more rows.")

    if "delivery_table_df" not in st.session_state:
        st.session_state.delivery_table_df = pd.DataFrame({
            "Item": [None] * 6, "Unit": [""] * 6, "Unit Cost": [0.0] * 6, "Quantity": [0.0] * 6
        })

    delivery_table_edited = st.data_editor(
        st.session_state.delivery_table_df,
        column_config={
            "Item": st.column_config.SelectboxColumn("Item", options=active_item_names_del, required=False, width="large"),
            "Unit": st.column_config.TextColumn("Unit", disabled=True, width="small"),
            "Unit Cost": st.column_config.NumberColumn("Unit Cost (₱)", disabled=True, format="₱%.2f", width="small"),
            "Quantity": st.column_config.NumberColumn("Quantity", min_value=0.0, step=0.01, format="%.2f", width="small"),
        },
        num_rows="dynamic",
        use_container_width=True,
        hide_index=True,
        key="delivery_table_editor",
    )

    # Keep Unit / Unit Cost in sync with whatever Item is picked on each row
    delivery_table_recomputed = delivery_table_edited.copy()
    delivery_table_recomputed["Unit"] = delivery_table_recomputed["Item"].map(item_unit_map_del).fillna("")
    delivery_table_recomputed["Unit Cost"] = delivery_table_recomputed["Item"].map(item_cost_map_del).fillna(0.0).astype(float).round(6)
    _current_display_del = delivery_table_edited[["Unit", "Unit Cost"]].fillna({"Unit": "", "Unit Cost": 0.0})
    _new_display_del = delivery_table_recomputed[["Unit", "Unit Cost"]]
    if not _current_display_del.reset_index(drop=True).equals(_new_display_del.reset_index(drop=True)):
        st.session_state.delivery_table_df = delivery_table_recomputed
        st.rerun()

    if st.button("🚀 Produce Delivery", type="primary", use_container_width=True):
        added, skipped = 0, 0
        for _, r in delivery_table_edited.iterrows():
            item_name = r.get("Item")
            qty = r.get("Quantity")
            if item_name and pd.notna(qty) and qty > 0:
                st.session_state.delivery_cart.append({
                    "item": item_name,
                    "qty": float(qty),
                    "unit": item_unit_map_del.get(item_name, ""),
                    "cost": item_cost_map_del.get(item_name, 0.0),
                    "notes": ""
                })
                added += 1
            elif item_name or (pd.notna(qty) and qty > 0):
                skipped += 1  # row half-filled (item without qty, or qty without item)

        if added == 0:
            st.warning("Fill in at least one row with both an item and a quantity.")
        else:
            st.session_state.delivery_table_df = pd.DataFrame({
                "Item": [None] * 6, "Unit": [""] * 6, "Unit Cost": [0.0] * 6, "Quantity": [0.0] * 6
            })
            msg = f"✅ {added} item(s) added to delivery."
            if skipped:
                msg += f" ({skipped} row(s) skipped — missing item or quantity.)"
            st.success(msg)
            st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)

    if st.session_state.delivery_cart:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown(f'<div class="section-title">Delivery Cart — {len(st.session_state.delivery_cart)} item(s)</div>', unsafe_allow_html=True)
        total_val = 0
        for idx, it in enumerate(st.session_state.delivery_cart):
            val = it["qty"] * it["cost"]
            total_val += val
            col_info, col_del = st.columns([5,1])
            with col_info:
                notes_str = f' &nbsp;·&nbsp; {it["notes"]}' if it['notes'] else ''
                st.markdown(f'<div class="po-item-row"><strong>{idx+1}. {it["item"]}</strong> &nbsp;|&nbsp; <span style="color:#8CAF7A;">{it["qty"]:,.2f} {it["unit"]}</span> &nbsp;|&nbsp; ₱{val:,.2f}{notes_str}</div>', unsafe_allow_html=True)
            with col_del:
                if st.button("🗑️", key=f"del_rm_{idx}"):
                    st.session_state.delivery_cart.pop(idx); st.rerun()

        st.markdown(f'<div class="info-box" style="text-align:right;">Total Delivery Value: <strong>₱{total_val:,.2f}</strong></div>', unsafe_allow_html=True)
        st.markdown("---")
        c1, c2 = st.columns([1,2])
        with c1:
            if st.button("🗑️ Clear", use_container_width=True):
                st.session_state.delivery_cart = []; st.rerun()
        with c2:
            if st.button("💾 Submit Delivery", type="primary", use_container_width=True):
                if not staff_name.strip():
                    st.error("Please enter your name.")
                else:
                    ref = f"DEL-{delivery_date.strftime('%Y%m%d')}-{datetime.now().strftime('%H%M%S')}"
                    log_ws = ensure_sheet(ss, LOG_SHEET, LOG_HEADERS)
                    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    rows = []
                    for it in st.session_state.delivery_cart:
                        rows.append([ts, delivery_date.strftime("%Y-%m-%d"),
                                     delivery_date.strftime("%b %Y").upper(),
                                     it["item"], staff_name.strip(), "DELIVERY", ref,
                                     it["qty"], 0, 0, 0, 0, 0, 0, 0, it["notes"], 0, 0])
                    log_ws.append_rows(rows)
                    invalidate_cache()
                    n = len(st.session_state.delivery_cart)
                    st.session_state.delivery_cart = []
                    st.success(f"✅ Delivery **{ref}** submitted! {n} item(s) logged.")
                    st.balloons()
        st.markdown('</div>', unsafe_allow_html=True)

    # ── Past Deliveries ──────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown('<div class="section-title">📂 Past Deliveries</div>', unsafe_allow_html=True)
    log_df = load_log(SPREADSHEET_ID)
    if not log_df.empty:
        del_log = log_df[log_df["TXN_TYPE"] == "DELIVERY"]
        if del_log.empty:
            st.info("No past deliveries yet.")
        else:
            past_refs = sorted(del_log["REF_NUMBER"].unique().tolist(), reverse=True)
            sel_ref = st.selectbox("Select Delivery Reference", past_refs, key="past_del_ref")
            if sel_ref:
                ref_rows = del_log[del_log["REF_NUMBER"] == sel_ref]
                first = ref_rows.iloc[0]
                total_val = sum(num(r["ADD_IN"]) * num(items_df[items_df["ITEM"]==r["ITEM"]]["UNIT COST"].values[0]) if r["ITEM"] in items_df["ITEM"].values else 0 for _, r in ref_rows.iterrows())
                st.markdown(f'<div class="info-box">📥 <strong>{sel_ref}</strong> &nbsp;|&nbsp; 📅 {first.get("DATE","")} &nbsp;|&nbsp; 👤 {first.get("STAFF","")} &nbsp;|&nbsp; {len(ref_rows)} item(s)</div>', unsafe_allow_html=True)

                for row_idx, (_, r) in enumerate(ref_rows.iterrows(), 1):
                    qty = num(r["ADD_IN"])
                    unit_cost = num(items_df[items_df["ITEM"]==r["ITEM"]]["UNIT COST"].values[0]) if r["ITEM"] in items_df["ITEM"].values else 0
                    uom = items_df[items_df["ITEM"]==r["ITEM"]]["UNIT OF MEASURE"].values[0] if r["ITEM"] in items_df["ITEM"].values else ""
                    val = qty * unit_cost
                    st.markdown(f'<div class="po-item-row"><strong>{r["ITEM"]}</strong> &nbsp;|&nbsp; <span style="color:#8CAF7A;">{qty:,.2f} {uom}</span> &nbsp;|&nbsp; ₱{unit_cost:.2f}/unit &nbsp;|&nbsp; ₱{val:,.2f}</div>', unsafe_allow_html=True)

                st.markdown(f'<div class="info-box" style="text-align:right;">Total Value: <strong>₱{total_val:,.2f}</strong></div>', unsafe_allow_html=True)

                del_rows = []
                for ri, (_, r) in enumerate(ref_rows.iterrows(), 1):
                    qty = num(r["ADD_IN"])
                    unit_cost = num(items_df[items_df["ITEM"]==r["ITEM"]]["UNIT COST"].values[0]) if r["ITEM"] in items_df["ITEM"].values else 0
                    uom = items_df[items_df["ITEM"]==r["ITEM"]]["UNIT OF MEASURE"].values[0] if r["ITEM"] in items_df["ITEM"].values else ""
                    del_rows.append((ri, r["ITEM"], uom, float(qty), qty * unit_cost))
                buf = build_doc_xlsx("INCOMING DELIVERY", sel_ref, first.get("DATE",""), first.get("STAFF",""), "Items", str(len(ref_rows)), del_rows, LOGO_B64)
                fname = f"{sel_ref}.xlsx"
                st.download_button(label="⬇️ Download Excel", data=buf, file_name=fname, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

                # ── Edit / Delete this Delivery ──────────────────────────────
                with st.expander("✏️ Edit or Delete this Delivery"):
                    if st.session_state.get("edit_del_ref") != sel_ref:
                        st.session_state.edit_del_ref = sel_ref
                        st.session_state.edit_del_cart = [
                            {"item": r["ITEM"], "qty": num(r["ADD_IN"])}
                            for _, r in ref_rows.iterrows()
                        ]

                    st.markdown(f"**Editing {sel_ref}** — adjust quantities, remove items, or add new ones, then save.")

                    for idx, it in enumerate(st.session_state.edit_del_cart):
                        c1, c2, c3 = st.columns([4,2,1])
                        with c1:
                            st.markdown(f'<div class="po-item-row">{it["item"]}</div>', unsafe_allow_html=True)
                        with c2:
                            new_qty = st.number_input("Qty", min_value=0.0, value=float(it["qty"]), step=0.01, format="%.2f", key=f"edit_del_qty_{sel_ref}_{idx}")
                            st.session_state.edit_del_cart[idx]["qty"] = new_qty
                        with c3:
                            if st.button("🗑️", key=f"edit_del_rm_{sel_ref}_{idx}"):
                                st.session_state.edit_del_cart.pop(idx); st.rerun()

                    st.markdown("---")
                    st.markdown("**Add an item to this delivery**")
                    ac1, ac2, ac3 = st.columns([3,1.5,1])
                    with ac1:
                        add_item_sel = st.selectbox("Item", active_items, key=f"edit_del_add_item_{sel_ref}")
                    with ac2:
                        add_item_qty = st.number_input("Qty", min_value=0.0, step=0.01, format="%.2f", key=f"edit_del_add_qty_{sel_ref}")
                    with ac3:
                        st.write("")
                        if st.button("➕ Add", key=f"edit_del_add_btn_{sel_ref}"):
                            add_item_name = item_name_map.get(add_item_sel, add_item_sel)
                            if add_item_qty > 0:
                                st.session_state.edit_del_cart.append({"item": add_item_name, "qty": add_item_qty})
                                st.rerun()

                    st.markdown("---")
                    confirm_save_del = st.checkbox("✅ I confirm these changes are correct", key=f"edit_del_confirm_save_{sel_ref}")
                    sc1, sc2 = st.columns(2)
                    with sc1:
                        if st.button("💾 Save Changes", type="primary", use_container_width=True, key=f"edit_del_save_{sel_ref}"):
                            if not confirm_save_del:
                                st.error("Please check the confirmation box before saving.")
                            else:
                                ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                new_rows = []
                                for it in st.session_state.edit_del_cart:
                                    if it["qty"] > 0:
                                        new_rows.append([ts, first.get("DATE",""), first.get("MONTH",""),
                                                          it["item"], first.get("STAFF",""), "DELIVERY", sel_ref,
                                                          it["qty"], 0, 0, 0, 0, 0, 0, 0, first.get("NOTES",""), 0, 0])
                                replace_ref_rows(ss, sel_ref, new_rows)
                                del st.session_state.edit_del_ref
                                del st.session_state.edit_del_cart
                                st.success(f"✅ {sel_ref} updated.")
                                st.rerun()
                    with sc2:
                        del_confirm = st.text_input("Type DELETE to remove this entire delivery", key=f"edit_del_confirm_{sel_ref}")
                        if st.button("💥 Delete Entire Delivery", use_container_width=True, key=f"edit_del_delete_{sel_ref}"):
                            if del_confirm.strip().upper() == "DELETE":
                                replace_ref_rows(ss, sel_ref, [])
                                if "edit_del_ref" in st.session_state: del st.session_state.edit_del_ref
                                if "edit_del_cart" in st.session_state: del st.session_state.edit_del_cart
                                st.success(f"✅ {sel_ref} deleted.")
                                st.rerun()
                            else:
                                st.error("Type DELETE exactly to confirm.")
    else:
        st.info("No past deliveries yet.")

# ══════════════════════════════════════════════════════════════════════════════
# PAGE: PURCHASE ORDERS
# ══════════════════════════════════════════════════════════════════════════════
elif page == "📋 Purchase Orders":
    st.markdown('<div class="main-header"><h1>📋 Purchase Orders</h1><p>Release stock to departments</p></div>', unsafe_allow_html=True)

    items_df = load_items(SPREADSHEET_ID)
    active_items = [f'{r["ITEM"]} ({r["UNIT OF MEASURE"]})' for _, r in items_df[items_df["ACTIVE"]=="YES"].iterrows()] if not items_df.empty else []
    item_name_map = {f'{r["ITEM"]} ({r["UNIT OF MEASURE"]})': r["ITEM"] for _, r in items_df[items_df["ACTIVE"]=="YES"].iterrows()} if not items_df.empty else {}

    if not active_items:
        st.warning("No items found. Go to Setup first.")
        st.stop()

    if "po_cart" not in st.session_state:
        st.session_state.po_cart = []

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">PO Details</div>', unsafe_allow_html=True)
    c1, c2, c3 = st.columns(3)
    with c1: po_date = st.date_input("📅 Date", value=date.today(), key="po_date")
    with c2: po_staff = st.text_input("👤 Prepared By", placeholder="Your name", key="po_staff")
    with c3:
        dept_choice = st.selectbox("🏢 Department", DEPARTMENTS + ["Others (Specify)"], key="po_dept")
    if dept_choice == "Others (Specify)":
        dept_specify = st.text_input("Specify department", key="po_dept_other")
        department = dept_specify.strip() if dept_specify.strip() else "Others"
    else:
        department = dept_choice
    st.markdown('</div>', unsafe_allow_html=True)

    active_item_names = items_df[items_df["ACTIVE"] == "YES"]["ITEM"].tolist()
    item_unit_map = {r["ITEM"]: r["UNIT OF MEASURE"] for _, r in items_df[items_df["ACTIVE"]=="YES"].iterrows()}
    item_cost_map = {r["ITEM"]: num(r["UNIT COST"]) for _, r in items_df[items_df["ACTIVE"]=="YES"].iterrows()}

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Add Items to PO</div>', unsafe_allow_html=True)
    st.caption(f"Pick an item and quantity per row. Use the **+** at the bottom of the table to add more rows. All rows go to the department selected above (**{department}**).")

    if "po_table_df" not in st.session_state:
        st.session_state.po_table_df = pd.DataFrame({
            "Item": [None] * 6, "Unit": [""] * 6, "Unit Cost": [0.0] * 6, "Quantity": [0.0] * 6
        })

    po_table_edited = st.data_editor(
        st.session_state.po_table_df,
        column_config={
            "Item": st.column_config.SelectboxColumn("Item", options=active_item_names, required=False, width="large"),
            "Unit": st.column_config.TextColumn("Unit", disabled=True, width="small"),
            "Unit Cost": st.column_config.NumberColumn("Unit Cost (₱)", disabled=True, format="₱%.2f", width="small"),
            "Quantity": st.column_config.NumberColumn("Quantity", min_value=0.0, step=0.01, format="%.2f", width="small"),
        },
        num_rows="dynamic",
        use_container_width=True,
        hide_index=True,
        key="po_table_editor",
    )

    # Keep Unit / Unit Cost in sync with whatever Item is picked on each row
    po_table_recomputed = po_table_edited.copy()
    po_table_recomputed["Unit"] = po_table_recomputed["Item"].map(item_unit_map).fillna("")
    po_table_recomputed["Unit Cost"] = po_table_recomputed["Item"].map(item_cost_map).fillna(0.0).astype(float).round(6)
    _current_display = po_table_edited[["Unit", "Unit Cost"]].fillna({"Unit": "", "Unit Cost": 0.0})
    _new_display = po_table_recomputed[["Unit", "Unit Cost"]]
    if not _current_display.reset_index(drop=True).equals(_new_display.reset_index(drop=True)):
        st.session_state.po_table_df = po_table_recomputed
        st.rerun()

    if st.button("🚀 Produce PO", type="primary", use_container_width=True):
        added, skipped = 0, 0
        for _, r in po_table_edited.iterrows():
            item_name = r.get("Item")
            qty = r.get("Quantity")
            if item_name and pd.notna(qty) and qty > 0:
                st.session_state.po_cart.append({
                    "item": item_name,
                    "qty": float(qty),
                    "notes": "",
                    "unit": item_unit_map.get(item_name, ""),
                    "cost": item_cost_map.get(item_name, 0.0),
                    "dept": department
                })
                added += 1
            elif item_name or (pd.notna(qty) and qty > 0):
                skipped += 1  # row half-filled (item without qty, or qty without item)

        if added == 0:
            st.warning("Fill in at least one row with both an item and a quantity.")
        else:
            st.session_state.po_table_df = pd.DataFrame({
                "Item": [None] * 6, "Unit": [""] * 6, "Unit Cost": [0.0] * 6, "Quantity": [0.0] * 6
            })
            msg = f"✅ {added} item(s) added to PO."
            if skipped:
                msg += f" ({skipped} row(s) skipped — missing item or quantity.)"
            st.success(msg)
            st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)

    if st.session_state.po_cart:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown(f'<div class="section-title">PO Review — {len(st.session_state.po_cart)} item(s) → {department}</div>', unsafe_allow_html=True)
        total_val = 0
        for idx, it in enumerate(st.session_state.po_cart):
            val = it["qty"] * it["cost"]
            total_val += val
            c_info, c_del = st.columns([5,1])
            with c_info:
                notes_str = f' &nbsp;·&nbsp; {it["notes"]}' if it['notes'] else ''
                st.markdown(f'<div class="po-item-row"><strong>{idx+1}. {it["item"]}</strong> &nbsp;|&nbsp; <span style="color:#8CAF7A;">{it["qty"]:,.2f} {it["unit"]}</span> &nbsp;|&nbsp; ₱{val:,.2f}{notes_str}</div>', unsafe_allow_html=True)
            with c_del:
                if st.button("🗑️", key=f"po_rm_{idx}"):
                    st.session_state.po_cart.pop(idx); st.rerun()

        st.markdown(f'<div class="info-box" style="text-align:right;">Total PO Value: <strong>₱{total_val:,.2f}</strong></div>', unsafe_allow_html=True)
        st.markdown("---")
        c1, c2 = st.columns([1,2])
        with c1:
            if st.button("🗑️ Clear PO", use_container_width=True):
                st.session_state.po_cart = []; st.rerun()
        with c2:
            if st.button("💾 Submit PO", type="primary", use_container_width=True):
                if not po_staff.strip():
                    st.error("Please enter your name.")
                else:
                    _dept_abbr = {"Resto":"RES","Café":"CAF","Bar":"BAR","Kitchen Ala Carte":"KAC","Kitchen Banquet":"KBQ","Others":"OTH"}
                    po_ref = f"PO-{po_date.strftime('%Y%m%d')}-{_dept_abbr.get(department, department[:3].upper())}-{datetime.now().strftime('%H%M%S')}"
                    log_ws = ensure_sheet(ss, LOG_SHEET, LOG_HEADERS)
                    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    rows = []
                    for it in st.session_state.po_cart:
                        dept_col = it["dept"]
                        r = [ts, po_date.strftime("%Y-%m-%d"), po_date.strftime("%b %Y").upper(),
                             it["item"], po_staff.strip(), "PO", po_ref,
                             0, 0, 0, 0, 0, 0, 0, 0, it["notes"], 0, 0]
                        if dept_col in DEPT_LOG_INDEX:
                            r[DEPT_LOG_INDEX[dept_col]] = it["qty"]
                        else:
                            r[13] = it["qty"]
                        rows.append(r)
                    log_ws.append_rows(rows)
                    invalidate_cache()
                    n = len(st.session_state.po_cart)
                    st.session_state.po_cart = []
                    st.success(f"✅ PO **{po_ref}** submitted! {n} item(s) released to **{department}**.")
                    st.balloons()
        st.markdown('</div>', unsafe_allow_html=True)

    # ── Past Purchase Orders ─────────────────────────────────────────────────
    st.markdown("---")
    st.markdown('<div class="section-title">📂 Past Purchase Orders</div>', unsafe_allow_html=True)
    log_df = load_log(SPREADSHEET_ID)
    if not log_df.empty:
        po_log = log_df[log_df["TXN_TYPE"] == "PO"]
        if po_log.empty:
            st.info("No past purchase orders yet.")
        else:
            past_po_refs = sorted(po_log["REF_NUMBER"].unique().tolist(), reverse=True)
            sel_po_ref = st.selectbox("Select PO Reference", past_po_refs, key="past_po_ref")
            if sel_po_ref:
                ref_rows = po_log[po_log["REF_NUMBER"] == sel_po_ref]
                first = ref_rows.iloc[0]
                dept_cols = ["RESTO","CAFE","BAR","KITCHEN_ALACARTE","KITCHEN_BANQUET","OTHERS"]
                def get_dept(r):
                    for dc in dept_cols:
                        if num(r.get(dc,0)) > 0:
                            return DEPT_DISPLAY.get(dc, dc.title()), num(r.get(dc,0))
                    return "Others", 0
                total_val = 0
                st.markdown(f'<div class="info-box">📋 <strong>{sel_po_ref}</strong> &nbsp;|&nbsp; 📅 {first.get("DATE","")} &nbsp;|&nbsp; 👤 {first.get("STAFF","")} &nbsp;|&nbsp; {len(ref_rows)} item(s)</div>', unsafe_allow_html=True)

                for row_idx, (_, r) in enumerate(ref_rows.iterrows(), 1):
                    dept_name, qty = get_dept(r)
                    unit_cost = num(items_df[items_df["ITEM"]==r["ITEM"]]["UNIT COST"].values[0]) if r["ITEM"] in items_df["ITEM"].values else 0
                    uom = items_df[items_df["ITEM"]==r["ITEM"]]["UNIT OF MEASURE"].values[0] if r["ITEM"] in items_df["ITEM"].values else ""
                    val = qty * unit_cost
                    total_val += val
                    st.markdown(f'<div class="po-item-row"><strong>{r["ITEM"]}</strong> &nbsp;|&nbsp; <span style="color:#8CAF7A;">{qty:,.2f} {uom}</span> &nbsp;|&nbsp; ₱{unit_cost:.2f}/unit &nbsp;|&nbsp; ₱{val:,.2f} &nbsp;|&nbsp; → {dept_name}</div>', unsafe_allow_html=True)

                st.markdown(f'<div class="info-box" style="text-align:right;">Total Value: <strong>₱{total_val:,.2f}</strong></div>', unsafe_allow_html=True)

                dept_display = sel_po_ref.split("-")[2].title() if len(sel_po_ref.split("-")) > 2 else "—"
                po_rows = []
                for ri, (_, r) in enumerate(ref_rows.iterrows(), 1):
                    dept_name, qty = get_dept(r)
                    unit_cost = num(items_df[items_df["ITEM"]==r["ITEM"]]["UNIT COST"].values[0]) if r["ITEM"] in items_df["ITEM"].values else 0
                    uom = items_df[items_df["ITEM"]==r["ITEM"]]["UNIT OF MEASURE"].values[0] if r["ITEM"] in items_df["ITEM"].values else ""
                    po_rows.append((ri, r["ITEM"], uom, float(qty), qty * unit_cost))
                buf = build_doc_xlsx("PURCHASE ORDER", sel_po_ref, first.get("DATE",""), first.get("STAFF",""), "Department", dept_display, po_rows, LOGO_B64)
                fname = f"{sel_po_ref}.xlsx"
                st.download_button(label="⬇️ Download Excel", data=buf, file_name=fname, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

                # ── Edit / Delete this PO ────────────────────────────────────
                with st.expander("✏️ Edit or Delete this Purchase Order"):
                    if st.session_state.get("edit_po_ref") != sel_po_ref:
                        st.session_state.edit_po_ref = sel_po_ref
                        _orig_dept = None
                        _cart = []
                        for _, r in ref_rows.iterrows():
                            dname, dqty = get_dept(r)
                            _orig_dept = dname
                            _cart.append({"item": r["ITEM"], "qty": dqty})
                        st.session_state.edit_po_cart = _cart
                        st.session_state.edit_po_dept = _orig_dept or "Resto"

                    st.markdown(f"**Editing {sel_po_ref}** — adjust quantities, remove items, add new ones, or change the department, then save.")

                    dept_idx = DEPARTMENTS.index(st.session_state.edit_po_dept) if st.session_state.edit_po_dept in DEPARTMENTS else 0
                    edit_po_dept_sel = st.selectbox("Department", DEPARTMENTS, index=dept_idx, key=f"edit_po_dept_sel_{sel_po_ref}")
                    st.session_state.edit_po_dept = edit_po_dept_sel

                    for idx, it in enumerate(st.session_state.edit_po_cart):
                        c1, c2, c3 = st.columns([4,2,1])
                        with c1:
                            st.markdown(f'<div class="po-item-row">{it["item"]}</div>', unsafe_allow_html=True)
                        with c2:
                            new_qty = st.number_input("Qty", min_value=0.0, value=float(it["qty"]), step=0.01, format="%.2f", key=f"edit_po_qty_{sel_po_ref}_{idx}")
                            st.session_state.edit_po_cart[idx]["qty"] = new_qty
                        with c3:
                            if st.button("🗑️", key=f"edit_po_rm_{sel_po_ref}_{idx}"):
                                st.session_state.edit_po_cart.pop(idx); st.rerun()

                    st.markdown("---")
                    st.markdown("**Add an item to this PO**")
                    ac1, ac2, ac3 = st.columns([3,1.5,1])
                    with ac1:
                        add_item_sel = st.selectbox("Item", active_items, key=f"edit_po_add_item_{sel_po_ref}")
                    with ac2:
                        add_item_qty = st.number_input("Qty", min_value=0.0, step=0.01, format="%.2f", key=f"edit_po_add_qty_{sel_po_ref}")
                    with ac3:
                        st.write("")
                        if st.button("➕ Add", key=f"edit_po_add_btn_{sel_po_ref}"):
                            add_item_name = item_name_map.get(add_item_sel, add_item_sel)
                            if add_item_qty > 0:
                                st.session_state.edit_po_cart.append({"item": add_item_name, "qty": add_item_qty})
                                st.rerun()

                    st.markdown("---")
                    confirm_save_po = st.checkbox("✅ I confirm these changes are correct", key=f"edit_po_confirm_save_{sel_po_ref}")
                    sc1, sc2 = st.columns(2)
                    with sc1:
                        if st.button("💾 Save Changes", type="primary", use_container_width=True, key=f"edit_po_save_{sel_po_ref}"):
                            if not confirm_save_po:
                                st.error("Please check the confirmation box before saving.")
                            else:
                                ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                new_rows = []
                                for it in st.session_state.edit_po_cart:
                                    if it["qty"] > 0:
                                        r = [ts, first.get("DATE",""), first.get("MONTH",""),
                                             it["item"], first.get("STAFF",""), "PO", sel_po_ref,
                                             0, 0, 0, 0, 0, 0, 0, 0, first.get("NOTES",""), 0, 0]
                                        dcol = DEPT_LOG_INDEX.get(st.session_state.edit_po_dept, 13)
                                        r[dcol] = it["qty"]
                                        new_rows.append(r)
                                replace_ref_rows(ss, sel_po_ref, new_rows)
                                del st.session_state.edit_po_ref
                                del st.session_state.edit_po_cart
                                del st.session_state.edit_po_dept
                                st.success(f"✅ {sel_po_ref} updated.")
                                st.rerun()
                    with sc2:
                        del_confirm = st.text_input("Type DELETE to remove this entire PO", key=f"edit_po_confirm_{sel_po_ref}")
                        if st.button("💥 Delete Entire PO", use_container_width=True, key=f"edit_po_delete_{sel_po_ref}"):
                            if del_confirm.strip().upper() == "DELETE":
                                replace_ref_rows(ss, sel_po_ref, [])
                                if "edit_po_ref" in st.session_state: del st.session_state.edit_po_ref
                                if "edit_po_cart" in st.session_state: del st.session_state.edit_po_cart
                                if "edit_po_dept" in st.session_state: del st.session_state.edit_po_dept
                                st.success(f"✅ {sel_po_ref} deleted.")
                                st.rerun()
                            else:
                                st.error("Type DELETE exactly to confirm.")
    else:
        st.info("No past purchase orders yet.")

# ══════════════════════════════════════════════════════════════════════════════
# PAGE: STOCK ADJUSTMENT
# ══════════════════════════════════════════════════════════════════════════════
elif page == "🔧 Stock Adjustment":
    st.markdown('<div class="main-header"><h1>🔧 Stock Adjustment</h1><p>Correct stock levels · Spoilage · Count adjustments</p></div>', unsafe_allow_html=True)

    items_df = load_items(SPREADSHEET_ID)
    active_items = [f'{r["ITEM"]} ({r["UNIT OF MEASURE"]})' for _, r in items_df[items_df["ACTIVE"]=="YES"].iterrows()] if not items_df.empty else []
    item_name_map = {f'{r["ITEM"]} ({r["UNIT OF MEASURE"]})': r["ITEM"] for _, r in items_df[items_df["ACTIVE"]=="YES"].iterrows()} if not items_df.empty else {}

    if not active_items:
        st.warning("No items found."); st.stop()

    if "adj_cart" not in st.session_state:
        st.session_state.adj_cart = []

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Adjustment Details</div>', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1: adj_date = st.date_input("📅 Date", value=date.today(), key="adj_date")
    with c2: adj_staff = st.text_input("👤 Adjusted By", placeholder="Your name", key="adj_staff")
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Add Adjustment</div>', unsafe_allow_html=True)
    c1, c2, c3 = st.columns([3,1.5,1])
    with c1: adj_item = st.selectbox("Select Item", active_items, key="adj_item")
    with c2:
        adj_type = st.selectbox("Type", ADJ_TYPES, key="adj_type")
    with c3: adj_qty = st.number_input("Qty", min_value=0.01, step=0.01, format="%.2f", key="adj_qty")
    adj_notes = st.text_input("Reason (required)", placeholder="e.g. physical count variance, expired goods...", key="adj_notes")

    if adj_item:
        _adj_item_name = item_name_map.get(adj_item, adj_item)
        _adj_matches = items_df[items_df["ITEM"]==_adj_item_name]
        if _adj_matches.empty:
            st.error(f"'{adj_item}' was not found in Items Master (it may have been renamed or removed). Please refresh the page or check the sheet.")
            st.stop()
        info = _adj_matches.iloc[0]
        st.markdown(f'<div class="info-box">📦 <strong>{adj_item}</strong> &nbsp;|&nbsp; {info["UNIT OF MEASURE"]} &nbsp;|&nbsp; ₱{num(info["UNIT COST"]):.2f}/unit</div>', unsafe_allow_html=True)

    if st.button("➕ Add Adjustment", use_container_width=True):
        if not adj_notes.strip():
            st.error("Please provide a reason for the adjustment.")
        elif adj_item and adj_qty > 0:
            st.session_state.adj_cart.append({
                "item": _adj_item_name, "type": adj_type, "qty": adj_qty,
                "notes": adj_notes, "unit": info["UNIT OF MEASURE"], "cost": num(info["UNIT COST"])
            })
            st.success(f"✅ {adj_item} added.")
            st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)

    if st.session_state.adj_cart:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown(f'<div class="section-title">Adjustments to Submit — {len(st.session_state.adj_cart)} item(s)</div>', unsafe_allow_html=True)
        for idx, it in enumerate(st.session_state.adj_cart):
            sign = "➕" if "Over" in it["type"] else "🛠️" if "Damaged" in it["type"] else "➖"
            c_i, c_d = st.columns([5,1])
            with c_i:
                st.markdown(f'<div class="po-item-row">{sign} <strong>{it["item"]}</strong> &nbsp;|&nbsp; {it["qty"]:,.2f} {it["unit"]} &nbsp;|&nbsp; <span style="color:#5A7A52;">{it["type"]}</span> &nbsp;|&nbsp; {it["notes"]}</div>', unsafe_allow_html=True)
            with c_d:
                if st.button("🗑️", key=f"adj_rm_{idx}"):
                    st.session_state.adj_cart.pop(idx); st.rerun()

        st.markdown("---")
        c1, c2 = st.columns([1,2])
        with c1:
            if st.button("🗑️ Clear", use_container_width=True):
                st.session_state.adj_cart = []; st.rerun()
        with c2:
            if st.button("💾 Submit Adjustments", type="primary", use_container_width=True):
                if not adj_staff.strip():
                    st.error("Please enter your name.")
                else:
                    ref = f"ADJ-{adj_date.strftime('%Y%m%d')}-{datetime.now().strftime('%H%M%S')}"
                    log_ws = ensure_sheet(ss, LOG_SHEET, LOG_HEADERS)
                    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    rows = []
                    for it in st.session_state.adj_cart:
                        over_qty   = it["qty"] if "Over" in it["type"] else 0
                        spoil_qty  = it["qty"] if "Spoilage" in it["type"] else 0
                        damage_qty = it["qty"] if "Damaged" in it["type"] else 0
                        rows.append([ts, adj_date.strftime("%Y-%m-%d"),
                                     adj_date.strftime("%b %Y").upper(),
                                     it["item"], adj_staff.strip(), "ADJUSTMENT", ref,
                                     0, over_qty, 0, 0, 0, 0, 0, spoil_qty, it["notes"], 0, damage_qty])
                    log_ws.append_rows(rows)
                    invalidate_cache()
                    n = len(st.session_state.adj_cart)
                    st.session_state.adj_cart = []
                    st.success(f"✅ Adjustment **{ref}** submitted! {n} item(s) updated.")
                    st.balloons()
        st.markdown('</div>', unsafe_allow_html=True)

    # ── Past Adjustments ─────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown('<div class="section-title">📂 Past Adjustments</div>', unsafe_allow_html=True)
    log_df = load_log(SPREADSHEET_ID)
    if not log_df.empty:
        adj_log = log_df[log_df["TXN_TYPE"] == "ADJUSTMENT"]
        if adj_log.empty:
            st.info("No past adjustments yet.")
        else:
            past_adj_refs = sorted(adj_log["REF_NUMBER"].unique().tolist(), reverse=True)
            sel_adj_ref = st.selectbox("Select Adjustment Reference", past_adj_refs, key="past_adj_ref")
            if sel_adj_ref:
                ref_rows = adj_log[adj_log["REF_NUMBER"] == sel_adj_ref]
                first = ref_rows.iloc[0]
                st.markdown(f'<div class="info-box">🔧 <strong>{sel_adj_ref}</strong> &nbsp;|&nbsp; 📅 {first.get("DATE","")} &nbsp;|&nbsp; 👤 {first.get("STAFF","")} &nbsp;|&nbsp; {len(ref_rows)} item(s)</div>', unsafe_allow_html=True)

                for _, r in ref_rows.iterrows():
                    over_v, spoil_v, dmg_v = num(r.get("OVER",0)), num(r.get("SPOILAGE",0)), num(r.get("DAMAGED_OUT",0))
                    if over_v > 0:
                        sign, qty_v, typ = "➕", over_v, "Over (Add +)"
                    elif dmg_v > 0:
                        sign, qty_v, typ = "🛠️", dmg_v, "Damaged (Remove −)"
                    else:
                        sign, qty_v, typ = "➖", spoil_v, "Spoilage (Remove −)"
                    notes_v = str(r.get("NOTES","")).strip()
                    st.markdown(f'<div class="po-item-row">{sign} <strong>{r["ITEM"]}</strong> &nbsp;|&nbsp; {qty_v:,.2f} &nbsp;|&nbsp; <span style="color:#5A7A52;">{typ}</span> &nbsp;|&nbsp; {notes_v}</div>', unsafe_allow_html=True)

                # ── Edit / Delete this Adjustment ────────────────────────────
                with st.expander("✏️ Edit or Delete this Adjustment"):
                    if st.session_state.get("edit_adj_ref") != sel_adj_ref:
                        st.session_state.edit_adj_ref = sel_adj_ref
                        _cart = []
                        for _, r in ref_rows.iterrows():
                            over_v, spoil_v, dmg_v = num(r.get("OVER",0)), num(r.get("SPOILAGE",0)), num(r.get("DAMAGED_OUT",0))
                            if over_v > 0:
                                _type, _qty = "Over (Add +)", over_v
                            elif dmg_v > 0:
                                _type, _qty = "Damaged (Remove −)", dmg_v
                            else:
                                _type, _qty = "Spoilage (Remove −)", spoil_v
                            _cart.append({
                                "item": r["ITEM"],
                                "type": _type,
                                "qty": _qty,
                                "notes": str(r.get("NOTES",""))
                            })
                        st.session_state.edit_adj_cart = _cart

                    st.markdown(f"**Editing {sel_adj_ref}** — adjust quantities/types, remove items, or add new ones, then save.")

                    for idx, it in enumerate(st.session_state.edit_adj_cart):
                        c1, c2, c3, c4 = st.columns([3,1.7,2,1])
                        with c1:
                            st.markdown(f'<div class="po-item-row">{it["item"]}</div>', unsafe_allow_html=True)
                        with c2:
                            new_type = st.selectbox("Type", ADJ_TYPES,
                                                     index=ADJ_TYPES.index(it["type"]) if it["type"] in ADJ_TYPES else 0,
                                                     key=f"edit_adj_type_{sel_adj_ref}_{idx}")
                            st.session_state.edit_adj_cart[idx]["type"] = new_type
                        with c3:
                            new_qty = st.number_input("Qty", min_value=0.0, value=float(it["qty"]), step=0.01, format="%.2f", key=f"edit_adj_qty_{sel_adj_ref}_{idx}")
                            st.session_state.edit_adj_cart[idx]["qty"] = new_qty
                        with c4:
                            if st.button("🗑️", key=f"edit_adj_rm_{sel_adj_ref}_{idx}"):
                                st.session_state.edit_adj_cart.pop(idx); st.rerun()
                        new_notes = st.text_input("Reason", value=it["notes"], key=f"edit_adj_notes_{sel_adj_ref}_{idx}")
                        st.session_state.edit_adj_cart[idx]["notes"] = new_notes

                    st.markdown("---")
                    st.markdown("**Add an item to this adjustment**")
                    ac1, ac2, ac3, ac4 = st.columns([3,1.7,1.3,1])
                    with ac1:
                        add_item_sel = st.selectbox("Item", active_items, key=f"edit_adj_add_item_{sel_adj_ref}")
                    with ac2:
                        add_item_type = st.selectbox("Type", ADJ_TYPES, key=f"edit_adj_add_type_{sel_adj_ref}")
                    with ac3:
                        add_item_qty = st.number_input("Qty", min_value=0.0, step=0.01, format="%.2f", key=f"edit_adj_add_qty_{sel_adj_ref}")
                    with ac4:
                        st.write("")
                        if st.button("➕ Add", key=f"edit_adj_add_btn_{sel_adj_ref}"):
                            add_item_name = item_name_map.get(add_item_sel, add_item_sel)
                            if add_item_qty > 0:
                                st.session_state.edit_adj_cart.append({"item": add_item_name, "type": add_item_type, "qty": add_item_qty, "notes": ""})
                                st.rerun()

                    st.markdown("---")
                    confirm_save_adj = st.checkbox("✅ I confirm these changes are correct", key=f"edit_adj_confirm_save_{sel_adj_ref}")
                    sc1, sc2 = st.columns(2)
                    with sc1:
                        if st.button("💾 Save Changes", type="primary", use_container_width=True, key=f"edit_adj_save_{sel_adj_ref}"):
                            if not confirm_save_adj:
                                st.error("Please check the confirmation box before saving.")
                            else:
                                ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                new_rows = []
                                for it in st.session_state.edit_adj_cart:
                                    if it["qty"] > 0:
                                        over_qty   = it["qty"] if "Over" in it["type"] else 0
                                        spoil_qty  = it["qty"] if "Spoilage" in it["type"] else 0
                                        damage_qty = it["qty"] if "Damaged" in it["type"] else 0
                                        new_rows.append([ts, first.get("DATE",""), first.get("MONTH",""),
                                                          it["item"], first.get("STAFF",""), "ADJUSTMENT", sel_adj_ref,
                                                          0, over_qty, 0, 0, 0, 0, 0, spoil_qty, it["notes"], 0, damage_qty])
                                replace_ref_rows(ss, sel_adj_ref, new_rows)
                                del st.session_state.edit_adj_ref
                                del st.session_state.edit_adj_cart
                                st.success(f"✅ {sel_adj_ref} updated.")
                                st.rerun()
                    with sc2:
                        del_confirm = st.text_input("Type DELETE to remove this entire adjustment", key=f"edit_adj_confirm_{sel_adj_ref}")
                        if st.button("💥 Delete Entire Adjustment", use_container_width=True, key=f"edit_adj_delete_{sel_adj_ref}"):
                            if del_confirm.strip().upper() == "DELETE":
                                replace_ref_rows(ss, sel_adj_ref, [])
                                if "edit_adj_ref" in st.session_state: del st.session_state.edit_adj_ref
                                if "edit_adj_cart" in st.session_state: del st.session_state.edit_adj_cart
                                st.success(f"✅ {sel_adj_ref} deleted.")
                                st.rerun()
                            else:
                                st.error("Type DELETE exactly to confirm.")
    else:
        st.info("No past adjustments yet.")
# ══════════════════════════════════════════════════════════════════════════════
# PAGE: ITEM HISTORY
# ══════════════════════════════════════════════════════════════════════════════
elif page == "🔍 Item History":
    st.markdown('<div class="main-header"><h1>🔍 Item History</h1><p>Full transaction log per item or per reference number</p></div>', unsafe_allow_html=True)

    items_df = load_items(SPREADSHEET_ID)
    log_df   = load_log(SPREADSHEET_ID)

    tab1, tab2 = st.tabs(["🔍 Search by Item", "🧾 Search by Reference (PO / Delivery / Adjustment)"])

    with tab1:
        if items_df.empty:
            st.info("No items yet.")
        else:
            search = st.text_input("Search Item", placeholder="Type item name...")
            active_names = items_df[items_df["ACTIVE"]=="YES"]["ITEM"].tolist()
            filtered_names = [i for i in active_names if search.lower() in i.lower()] if search else active_names
            sel = st.selectbox("Select Item", filtered_names, key="hist_item")

            if sel:
                info = items_df[items_df["ITEM"]==sel].iloc[0]
                st.markdown(f'<div class="info-box">📦 <strong>{sel}</strong> &nbsp;|&nbsp; {info["UNIT OF MEASURE"]} &nbsp;|&nbsp; ₱{num(info["UNIT COST"]):.2f}/unit &nbsp;|&nbsp; {info["CATEGORY"].upper()} &nbsp;|&nbsp; Current Beginning: <strong>{num(info["BEGINNING_STOCKS"]):,.2f}</strong></div>', unsafe_allow_html=True)

                if log_df.empty:
                    st.info("No transactions yet.")
                else:
                    ilog = log_df[log_df["ITEM"]==sel].copy()
                    if ilog.empty:
                        st.info("No transactions recorded for this item yet.")
                    else:
                        total_in    = ilog["ADD_IN"].apply(num).sum()
                        total_over  = ilog["OVER"].apply(num).sum()
                        total_out   = (ilog["RESTO"].apply(num) + ilog["KITCHEN_BANQUET"].apply(num) +
                                       ilog["KITCHEN_ALACARTE"].apply(num) +
                                       ilog["CAFE"].apply(num) + ilog["BAR"].apply(num) +
                                       ilog["OTHERS"].apply(num)).sum()
                        total_spoil  = ilog["SPOILAGE"].apply(num).sum()
                        total_damage = ilog["DAMAGED_OUT"].apply(num).sum()
                        beginning   = num(info["BEGINNING_STOCKS"])
                        ending_stock = beginning + total_in + total_over - total_out - total_spoil - total_damage

                        c1,c2,c3,c4,c5,c6 = st.columns(6)
                        with c1: st.metric("Total Incoming", f"{total_in:,.2f}")
                        with c2: st.metric("Total Released (PO)", f"{total_out:,.2f}")
                        with c3: st.metric("Total Spoilage", f"{total_spoil:,.2f}")
                        with c4: st.metric("Total Damaged", f"{total_damage:,.2f}")
                        with c5: st.metric("Ending Stock", f"{ending_stock:,.2f}")
                        with c6: st.metric("Total Transactions", len(ilog))

                        st.markdown("---")

                        for txn_type, icon, color in [
                            ("DELIVERY",   "📥", "#4A8ACC"),
                            ("PO",         "📋", "#8CAF7A"),
                            ("ADJUSTMENT", "🔧", "#C4A840"),
                            ("CARRYOVER",  "🔄", "#AA6ACC"),
                        ]:
                            tlog = ilog[ilog["TXN_TYPE"]==txn_type].sort_values("TIMESTAMP", ascending=False)
                            if tlog.empty:
                                continue

                            st.markdown(f'<div style="font-size:0.65rem;font-weight:600;letter-spacing:2px;text-transform:uppercase;color:{color};margin:12px 0 6px 0;padding-bottom:4px;border-bottom:1px solid #1E2E1C;">{icon} {txn_type} — {len(tlog)} record(s)</div>', unsafe_allow_html=True)

                            for _, row in tlog.iterrows():
                                ref  = row.get("REF_NUMBER","")
                                date_str = row.get("DATE","")
                                staff = row.get("STAFF","")
                                parts = []
                                for col, label in [("ADD_IN","Incoming"),("OVER","Over (Adjustment)"),
                                                   ("RESTO","Resto"),("KITCHEN_BANQUET","Kitchen Banquet"),
                                                   ("KITCHEN_ALACARTE","Kitchen Ala Carte"),
                                                   ("CAFE","Café"),("BAR","Bar"),
                                                   ("OTHERS","Others"),("SPOILAGE","Spoilage"),("DAMAGED_OUT","Damaged")]:
                                    v = num(row.get(col,0))
                                    if v: parts.append(f"<strong style='color:{color};'>{label}:</strong> {v:,.2f}")
                                _notes_val = str(row.get("NOTES", "")).strip()
                                notes_str = f'<div style="color:#5A7A52;font-size:0.78rem;margin-top:3px;">📝 {_notes_val}</div>' if _notes_val else ""
                                st.markdown(f'<div class="log-entry" style="border-left-color:{color};"><div style="display:flex;justify-content:space-between;align-items:center;"><span>📅 <strong>{date_str}</strong> &nbsp;·&nbsp; 👤 {staff}</span><span style="color:#3A5238;font-size:0.75rem;">{ref}</span></div><div style="margin-top:5px;">{" &nbsp;&nbsp; ".join(parts) if parts else "—"}</div>{notes_str}</div>', unsafe_allow_html=True)

    with tab2:
        ref_search = st.text_input("Search Reference #", placeholder="e.g. PO-20260628, DEL-..., ADJ-...")
        if ref_search and not log_df.empty:
            rlog = log_df[log_df["REF_NUMBER"].astype(str).str.contains(ref_search, case=False, na=False)]
            if rlog.empty:
                st.warning("No records found.")
            else:
                for ref in rlog["REF_NUMBER"].unique():
                    rr = rlog[rlog["REF_NUMBER"]==ref]
                    first = rr.iloc[0]
                    st.markdown(f'<div class="info-box">🧾 <strong>{ref}</strong> &nbsp;|&nbsp; {first.get("TXN_TYPE","")} &nbsp;|&nbsp; 📅 {first.get("DATE","")} &nbsp;|&nbsp; 👤 {first.get("STAFF","")} &nbsp;|&nbsp; {len(rr)} item(s)</div>', unsafe_allow_html=True)
                    for _, row in rr.iterrows():
                        parts = []
                        for col, label in [("ADD_IN","Incoming"),("OVER","Over"),
                                           ("RESTO","Resto"),("KITCHEN_BANQUET","Kitchen Banquet"),
                                           ("KITCHEN_ALACARTE","Kitchen Ala Carte"),
                                           ("CAFE","Café"),("BAR","Bar"),("OTHERS","Others"),
                                           ("SPOILAGE","Spoilage"),("DAMAGED_OUT","Damaged")]:
                            v = num(row.get(col,0))
                            if v: parts.append(f"{label}: <strong>{v:,.2f}</strong>")
                        st.markdown(f'<div class="log-entry">📦 <strong>{row.get("ITEM","")}</strong><div style="margin-top:3px;">{" &nbsp;·&nbsp; ".join(parts) if parts else "—"}</div></div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# PAGE: ITEMS MASTER
# ══════════════════════════════════════════════════════════════════════════════
elif page == "📦 Items Master":
    st.markdown('<div class="main-header"><h1>📦 Items Master</h1><p>Manage your ingredient list</p></div>', unsafe_allow_html=True)

    items_df = load_items(SPREADSHEET_ID)
    tab1, tab2, tab3 = st.tabs(["➕ Add New Item", "✏️ Edit / Deactivate", "📋 View All"])

    with tab1:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        with c1:
            new_name = st.text_input("Item Name *")
            new_unit = st.text_input("Unit of Measure *", placeholder="gram / ml / piece / bottle")
        with c2:
            new_cost = st.number_input("Unit Cost (₱) *", min_value=0.0, step=0.0001, format="%.4f")
            new_cat  = st.selectbox("Category *", CATEGORIES)
        new_begin = st.number_input("Beginning Stock", min_value=0.0, step=0.01, format="%.2f")

        if st.button("➕ Add Item", type="primary"):
            if not new_name.strip() or not new_unit.strip():
                st.error("Name and unit are required.")
            elif not items_df.empty and new_name.strip() in items_df["ITEM"].values:
                st.error("Item already exists.")
            else:
                ws = ensure_sheet(ss, ITEMS_SHEET, ITEMS_HEADERS)
                ws.append_row([new_name.strip(), new_cost, new_unit.strip(), new_cat, new_begin, "YES"])
                invalidate_cache()
                st.success(f"✅ {new_name} added!"); st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    with tab2:
        if items_df.empty:
            st.info("No items yet.")
        else:
            search = st.text_input("🔍 Search", key="edit_search")
            fdf = items_df[items_df["ITEM"].str.contains(search, case=False, na=False)] if search else items_df
            sel = st.selectbox("Select Item", fdf["ITEM"].tolist())
            if sel:
                row = items_df[items_df["ITEM"]==sel].iloc[0]
                ridx = items_df[items_df["ITEM"]==sel].index[0] + 2
                st.markdown('<div class="card">', unsafe_allow_html=True)
                c1, c2 = st.columns(2)
                with c1:
                    e_cost  = st.number_input("Unit Cost", value=num(row["UNIT COST"]), step=0.0001, format="%.4f", key="e_cost")
                    e_unit  = st.text_input("Unit of Measure", value=str(row["UNIT OF MEASURE"]), key="e_unit")
                    e_begin = st.number_input("Beginning Stock", value=num(row["BEGINNING_STOCKS"]), step=0.01, format="%.2f", key="e_begin")
                with c2:
                    e_cat    = st.selectbox("Category", CATEGORIES, index=CATEGORIES.index(row["CATEGORY"]) if row["CATEGORY"] in CATEGORIES else 0, key="e_cat")
                    e_active = st.selectbox("Status", ["YES","NO"], index=0 if row["ACTIVE"]=="YES" else 1, key="e_active")
                if st.button("💾 Save Changes", type="primary"):
                    ws = ss.worksheet(ITEMS_SHEET)
                    ws.update_cell(ridx, 2, e_cost)
                    ws.update_cell(ridx, 3, e_unit)
                    ws.update_cell(ridx, 4, e_cat)
                    ws.update_cell(ridx, 5, e_begin)
                    ws.update_cell(ridx, 6, e_active)
                    invalidate_cache()
                    st.success(f"✅ {sel} updated!"); st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)

    with tab3:
        if items_df.empty:
            st.info("No items yet.")
        else:
            show_all = st.checkbox("Show inactive items")
            df = items_df if show_all else items_df[items_df["ACTIVE"]=="YES"]
            st.dataframe(df, use_container_width=True, hide_index=True)
            st.caption(f"{len(df)} items")

# ══════════════════════════════════════════════════════════════════════════════
# PAGE: EXPORT TO EXCEL
# ══════════════════════════════════════════════════════════════════════════════
elif page == "⬇️ Export to Excel":
    st.markdown('<div class="main-header"><h1>⬇️ Export to Excel</h1><p>Generate inventory report for any date range</p></div>', unsafe_allow_html=True)

    items_df = load_items(SPREADSHEET_ID)
    log_df   = load_log(SPREADSHEET_ID)

    if items_df.empty:
        st.warning("No items yet. Go to Setup first."); st.stop()

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Select Date Range</div>', unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    with c1:
        from_date = st.date_input("📅 From Date", value=date.today().replace(day=1), key="exp_from")
    with c2:
        to_date = st.date_input("📅 To Date", value=date.today(), key="exp_to")

    if from_date > to_date:
        st.error("From date must be before To date.")
        st.stop()

    from_str = from_date.strftime("%Y-%m-%d")
    to_str   = to_date.strftime("%Y-%m-%d")

    if not log_df.empty and "DATE" in log_df.columns:
        range_log = log_df[(log_df["DATE"] >= from_str) & (log_df["DATE"] <= to_str)]
        st.markdown(f'<div class="info-box">📊 <strong>{len(range_log)}</strong> transactions found between <strong>{from_date.strftime("%b %d, %Y")}</strong> and <strong>{to_date.strftime("%b %d, %Y")}</strong></div>', unsafe_allow_html=True)
    else:
        range_log = pd.DataFrame()
        st.markdown('<div class="info-box">No transactions found in this range.</div>', unsafe_allow_html=True)

    if st.button("📥 Generate Excel File", type="primary", use_container_width=True):
        from datetime import timedelta
        all_days = []
        cur = from_date
        while cur <= to_date:
            all_days.append(cur.strftime("%Y-%m-%d"))
            cur += timedelta(days=1)

        if from_date.month == to_date.month and from_date.year == to_date.year:
            file_label = from_date.strftime("%B_%Y")
            title_label = from_date.strftime("%B %Y").upper()
        else:
            file_label = f"{from_date.strftime('%b_%Y')}_to_{to_date.strftime('%b_%Y')}"
            title_label = f"{from_date.strftime('%b %d').upper()} TO {to_date.strftime('%b %d, %Y').upper()}"

        active = items_df[items_df["ACTIVE"] == "YES"].copy().reset_index(drop=True)

        wb = openpyxl.Workbook()

        hdr_fill   = PatternFill("solid", fgColor="1A2E1A")
        title_fill = PatternFill("solid", fgColor="0A1208")
        worth_fill = PatternFill("solid", fgColor="0E1C0E")
        hdr_font   = Font(bold=True, color="A8C896", size=9)
        title_font = Font(bold=True, color="C8DCC0", size=12)
        worth_font = Font(bold=True, color="8CAF7A", size=9)
        thin       = Side(style="thin", color="2A3828")
        border     = Border(left=thin, right=thin, top=thin, bottom=thin)
        num_fmt4   = "#,##0.00"

        cat_colors = {
            "beverage":"D6E8F5","beef":"F5D6D6","chicken":"FFF3D6",
            "seafood":"D6F0F5","fresh":"D6F5E0","dry":"FFF8D6",
            "wet":"EDD6F5","rtc":"D6E8FF","pork":"F5D6D6",
            "meat":"F0F0F0","frozen":"D6EAF8","dessert":"FFF0D6"
        }

        MAIN_COLS  = ["ITEM","BEGINNING STOCKS","UNIT COST","UNIT OF MEASURE","CATEGORY",
                      "ADD'L / IN","OVER","DAMAGED / OUT","CAFÉ","BAR","KITCHEN ALA CARTE","KITCHEN BANQUET","RESTO","OTHERS",
                      "SPOILAGE","ENDING STOCKS","TOTAL WORTH OF STOCKS"]
        WORTH_COLS = ["WORTH OF BEGINNING","WORTH OF ADD'L/IN","WORTH OF OVER","WORTH OF DAMAGED",
                      "WORTH OF CAFÉ","WORTH OF BAR","WORTH OF KITCHEN ALA CARTE","WORTH OF KITCHEN BANQUET",
                      "WORTH OF RESTO","WORTH OF OTHERS","WORTH OF SPOILAGE","WORTH OF ENDING"]
        MAIN_NUM_COLS  = [2,3,6,7,8,9,10,11,12,13,14,15,16,17]   # all numeric main columns (incl. Unit Cost)
        MAIN_SUM_COLS  = [2,6,7,8,9,10,11,12,13,14,15,16,17]     # numeric main columns to total (excl. Unit Cost)
        WORTH_START_COL = 19  # main block (17 cols) + 1 blank spacer column (18)

        def col_letter(n): return get_column_letter(n)

        item_running = {row["ITEM"]: num(row["BEGINNING_STOCKS"]) for _, row in active.iterrows()}

        def write_day_sheet(ws_out, day_str, day_label, tab_log, item_running):
            ws_out.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(MAIN_COLS))
            tc = ws_out["A1"]
            tc.value = f"{day_label} — SERVANDO MAIN WAREHOUSE INVENTORY"
            tc.font = title_font; tc.fill = title_fill
            tc.alignment = Alignment(horizontal="center", vertical="center")
            ws_out.row_dimensions[1].height = 26

            for ci, col in enumerate(MAIN_COLS, 1):
                c = ws_out.cell(row=2, column=ci, value=col)
                c.fill = hdr_fill; c.font = hdr_font
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = border
            for ci, col in enumerate(WORTH_COLS, WORTH_START_COL):
                c = ws_out.cell(row=2, column=ci, value=col)
                c.fill = worth_fill; c.font = worth_font
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = border
            ws_out.row_dimensions[2].height = 38

            for ri, (_, item) in enumerate(active.iterrows(), 3):
                name = item["ITEM"]
                cost = num(item["UNIT COST"])
                cat  = str(item["CATEGORY"]).lower()
                beg  = item_running.get(name, 0.0)

                if not tab_log.empty:
                    ilog   = tab_log[tab_log["ITEM"] == name]
                    add_in = ilog["ADD_IN"].apply(num).sum()
                    over   = ilog["OVER"].apply(num).sum()
                    dmg    = ilog["DAMAGED_OUT"].apply(num).sum()
                    cafe   = ilog["CAFE"].apply(num).sum()
                    bar    = ilog["BAR"].apply(num).sum()
                    kac    = ilog["KITCHEN_ALACARTE"].apply(num).sum()
                    kbq    = ilog["KITCHEN_BANQUET"].apply(num).sum()
                    resto  = ilog["RESTO"].apply(num).sum()
                    others = ilog["OTHERS"].apply(num).sum()
                    spoil  = ilog["SPOILAGE"].apply(num).sum()
                else:
                    add_in=over=dmg=cafe=bar=kac=kbq=resto=others=spoil=0.0

                ending = beg + add_in + over - dmg - cafe - bar - kac - kbq - resto - others - spoil
                worth  = ending * cost
                cfill  = PatternFill("solid", fgColor=cat_colors.get(cat, "FFFFFF"))

                main_vals = [name, round(beg,4), round(cost,6), item["UNIT OF MEASURE"],
                             item["CATEGORY"], round(add_in,4), round(over,4), round(dmg,4),
                             round(cafe,4), round(bar,4), round(kac,4), round(kbq,4), round(resto,4),
                             round(others,4), round(spoil,4), round(ending,4), round(worth,4)]
                worth_vals = [beg*cost, add_in*cost, over*cost, dmg*cost, cafe*cost, bar*cost,
                              kac*cost, kbq*cost, resto*cost, others*cost, spoil*cost, worth]

                for ci, val in enumerate(main_vals, 1):
                    cell = ws_out.cell(row=ri, column=ci, value=val)
                    cell.fill = cfill; cell.border = border
                    if ci in MAIN_NUM_COLS:
                        cell.number_format = num_fmt4
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = Alignment(vertical="center")

                for ci, val in enumerate(worth_vals, WORTH_START_COL):
                    cell = ws_out.cell(row=ri, column=ci, value=round(val,4))
                    cell.fill = worth_fill
                    cell.font = Font(color="8CAF7A", size=9)
                    cell.number_format = num_fmt4; cell.border = border
                    cell.alignment = Alignment(horizontal="right", vertical="center")

            total_row = 3 + len(active) + 20
            ws_out.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, color="FFFFFF")
            ws_out.cell(row=total_row, column=1).fill = hdr_fill
            ws_out.cell(row=total_row, column=1).border = border
            for ci in MAIN_SUM_COLS:
                col_l = col_letter(ci)
                cell = ws_out.cell(row=total_row, column=ci,
                    value=f"=SUM({col_l}3:{col_l}{total_row-1})")
                cell.fill = hdr_fill; cell.font = Font(bold=True, color="A8C896", size=9)
                cell.number_format = num_fmt4; cell.border = border
                cell.alignment = Alignment(horizontal="right")
            for ci in range(WORTH_START_COL, WORTH_START_COL+len(WORTH_COLS)):
                col_l = col_letter(ci)
                cell = ws_out.cell(row=total_row, column=ci,
                    value=f"=SUM({col_l}3:{col_l}{total_row-1})")
                cell.fill = worth_fill; cell.font = Font(bold=True, color="8CAF7A", size=9)
                cell.number_format = num_fmt4; cell.border = border
                cell.alignment = Alignment(horizontal="right")

            col_widths = [32,13,11,14,11,11,10,12,10,10,16,16,11,10,12,14,17,3,
                          14,13,12,13,12,12,18,18,13,13,14,14]
            for ci, w in enumerate(col_widths, 1):
                ws_out.column_dimensions[col_letter(ci)].width = w
            ws_out.freeze_panes = "A3"

            for _, item in active.iterrows():
                name = item["ITEM"]
                beg  = item_running.get(name, 0.0)
                if not tab_log.empty:
                    ilog   = tab_log[tab_log["ITEM"] == name]
                    add_in = ilog["ADD_IN"].apply(num).sum()
                    over   = ilog["OVER"].apply(num).sum()
                    dmg    = ilog["DAMAGED_OUT"].apply(num).sum()
                    cafe   = ilog["CAFE"].apply(num).sum()
                    bar    = ilog["BAR"].apply(num).sum()
                    kac    = ilog["KITCHEN_ALACARTE"].apply(num).sum()
                    kbq    = ilog["KITCHEN_BANQUET"].apply(num).sum()
                    resto  = ilog["RESTO"].apply(num).sum()
                    others = ilog["OTHERS"].apply(num).sum()
                    spoil  = ilog["SPOILAGE"].apply(num).sum()
                else:
                    add_in=over=dmg=cafe=bar=kac=kbq=resto=others=spoil=0.0
                item_running[name] = beg + add_in + over - dmg - cafe - bar - kac - kbq - resto - others - spoil
            return item_running

        for day_str in all_days:
            try:
                d = datetime.strptime(day_str, "%Y-%m-%d")
                tab_name  = str(d.day)
                day_label = d.strftime("%B %d, %Y").upper()
            except:
                tab_name  = day_str[-2:].lstrip("0") or "1"
                day_label = day_str

            tab_log = range_log[range_log["DATE"] == day_str] if not range_log.empty else pd.DataFrame()
            ws_day  = wb.create_sheet(title=tab_name)
            item_running = write_day_sheet(ws_day, day_str, day_label, tab_log, item_running)

        ws_sum = wb.create_sheet(title="SUMMARY")
        ws_sum.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(MAIN_COLS))
        tc = ws_sum["A1"]
        tc.value = f"SUMMARY — {title_label} — SERVANDO MAIN WAREHOUSE INVENTORY"
        tc.font = title_font; tc.fill = title_fill
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws_sum.row_dimensions[1].height = 26

        for ci, col in enumerate(MAIN_COLS, 1):
            c = ws_sum.cell(row=2, column=ci, value=col)
            c.fill = hdr_fill; c.font = hdr_font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
        for ci, col in enumerate(WORTH_COLS, WORTH_START_COL):
            c = ws_sum.cell(row=2, column=ci, value=col)
            c.fill = worth_fill; c.font = worth_font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
        ws_sum.row_dimensions[2].height = 38

        for ri, (_, item) in enumerate(active.iterrows(), 3):
            name = item["ITEM"]
            cost = num(item["UNIT COST"])
            cat  = str(item["CATEGORY"]).lower()
            beg  = num(item["BEGINNING_STOCKS"])

            if not range_log.empty:
                ilog   = range_log[range_log["ITEM"] == name]
                add_in = ilog["ADD_IN"].apply(num).sum()
                over   = ilog["OVER"].apply(num).sum()
                dmg    = ilog["DAMAGED_OUT"].apply(num).sum()
                cafe   = ilog["CAFE"].apply(num).sum()
                bar    = ilog["BAR"].apply(num).sum()
                kac    = ilog["KITCHEN_ALACARTE"].apply(num).sum()
                kbq    = ilog["KITCHEN_BANQUET"].apply(num).sum()
                resto  = ilog["RESTO"].apply(num).sum()
                others = ilog["OTHERS"].apply(num).sum()
                spoil  = ilog["SPOILAGE"].apply(num).sum()
            else:
                add_in=over=dmg=cafe=bar=kac=kbq=resto=others=spoil=0.0

            ending = beg + add_in + over - dmg - cafe - bar - kac - kbq - resto - others - spoil
            worth  = ending * cost
            cfill  = PatternFill("solid", fgColor=cat_colors.get(cat, "FFFFFF"))

            main_vals = [name, round(beg,4), round(cost,6), item["UNIT OF MEASURE"],
                         item["CATEGORY"], round(add_in,4), round(over,4), round(dmg,4),
                         round(cafe,4), round(bar,4), round(kac,4), round(kbq,4), round(resto,4),
                         round(others,4), round(spoil,4), round(ending,4), round(worth,4)]
            worth_vals = [beg*cost, add_in*cost, over*cost, dmg*cost, cafe*cost, bar*cost,
                          kac*cost, kbq*cost, resto*cost, others*cost, spoil*cost, worth]

            for ci, val in enumerate(main_vals, 1):
                cell = ws_sum.cell(row=ri, column=ci, value=val)
                cell.fill = cfill; cell.border = border
                if ci in MAIN_NUM_COLS:
                    cell.number_format = num_fmt4
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(vertical="center")
            for ci, val in enumerate(worth_vals, WORTH_START_COL):
                cell = ws_sum.cell(row=ri, column=ci, value=round(val,4))
                cell.fill = worth_fill
                cell.font = Font(color="8CAF7A", size=9)
                cell.number_format = num_fmt4; cell.border = border
                cell.alignment = Alignment(horizontal="right", vertical="center")

        total_row = 3 + len(active) + 20
        ws_sum.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, color="FFFFFF")
        ws_sum.cell(row=total_row, column=1).fill = hdr_fill
        ws_sum.cell(row=total_row, column=1).border = border
        for ci in MAIN_SUM_COLS:
            col_l = col_letter(ci)
            cell = ws_sum.cell(row=total_row, column=ci, value=f"=SUM({col_l}3:{col_l}{total_row-1})")
            cell.fill = hdr_fill; cell.font = Font(bold=True, color="A8C896", size=9)
            cell.number_format = num_fmt4; cell.border = border
            cell.alignment = Alignment(horizontal="right")
        for ci in range(WORTH_START_COL, WORTH_START_COL+len(WORTH_COLS)):
            col_l = col_letter(ci)
            cell = ws_sum.cell(row=total_row, column=ci, value=f"=SUM({col_l}3:{col_l}{total_row-1})")
            cell.fill = worth_fill; cell.font = Font(bold=True, color="8CAF7A", size=9)
            cell.number_format = num_fmt4; cell.border = border
            cell.alignment = Alignment(horizontal="right")

        col_widths = [32,13,11,14,11,11,10,12,10,10,16,16,11,10,12,14,17,3,
                      14,13,12,13,12,12,18,18,13,13,14,14]
        for ci, w in enumerate(col_widths, 1):
            ws_sum.column_dimensions[col_letter(ci)].width = w
        ws_sum.freeze_panes = "A3"

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        buf = BytesIO()
        wb.save(buf); buf.seek(0)

        st.download_button(
            label=f"📥 Download {file_label.replace('_', ' ')}.xlsx",
            data=buf,
            file_name=f"{file_label}_-_Servando_Main_Warehouse.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.success(f"✅ Done! {len(all_days)} daily tabs + SUMMARY. File: {file_label} - Servando Main Warehouse.xlsx")
    st.markdown('</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# PAGE: SETUP
# ══════════════════════════════════════════════════════════════════════════════
elif page == "⚙️ Setup":
    st.markdown('<div class="main-header"><h1>⚙️ Setup</h1><p>Initialize system · Import items</p></div>', unsafe_allow_html=True)

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Current Status</div>', unsafe_allow_html=True)
    items_df = load_items(SPREADSHEET_ID)
    log_df   = load_log(SPREADSHEET_ID)
    c1, c2, c3 = st.columns(3)
    with c1: st.metric("Items in Master", len(items_df))
    with c2: st.metric("Total Transactions", len(log_df))
    with c3: st.metric("Active Items", len(items_df[items_df["ACTIVE"]=="YES"]) if not items_df.empty else 0)
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown("---")
    st.markdown('<div class="section-title" style="color:#CC6A6A;">⚠️ Danger Zone — Reset / Clear Data</div>', unsafe_allow_html=True)
    st.markdown('<div style="color:#CC6A6A;font-size:0.82rem;margin-bottom:1rem;">These actions are permanent and cannot be undone. Type the confirmation word before proceeding.</div>', unsafe_allow_html=True)

    # Clear Daily Transactions
    st.markdown('<div class="card" style="border-color:#3A1A1A;">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Clear Daily Transactions</div>', unsafe_allow_html=True)
    st.write("Wipes Incoming Deliveries, Purchase Orders, and Stock Adjustments. Your Item List and Beginning Inventory stay untouched.")
    confirm1 = st.text_input("Type **CLEAR TRANSACTIONS** to confirm", key="confirm_txn")
    if st.button("🗑️ Clear Daily Transactions", key="btn_clear_txn"):
        if confirm1.strip().upper() == "CLEAR TRANSACTIONS":
            ws = ensure_sheet(ss, LOG_SHEET, LOG_HEADERS)
            ws.clear()
            ws.append_row(LOG_HEADERS)
            invalidate_cache()
            st.success("✅ Daily transactions cleared. Item List is untouched.")
        else:
            st.error("❌ Confirmation text doesn't match. Nothing was deleted.")
    st.markdown('</div>', unsafe_allow_html=True)

    # Clear Item List
    st.markdown('<div class="card" style="border-color:#5A1A1A;">', unsafe_allow_html=True)
    st.markdown('<div class="section-title" style="color:#CC6A6A;">Clear Item List</div>', unsafe_allow_html=True)
    st.write("Wipes the entire Items List and Beginning Inventory. Daily transactions (deliveries, POs, adjustments) are kept.")
    confirm2 = st.text_input("Type **CLEAR ITEMS** to confirm", key="confirm_items")
    if st.button("🗑️ Clear Item List", key="btn_clear_items"):
        if confirm2.strip().upper() == "CLEAR ITEMS":
            ws = ensure_sheet(ss, ITEMS_SHEET, ITEMS_HEADERS)
            ws.clear()
            ws.append_row(ITEMS_HEADERS)
            invalidate_cache()
            st.success("✅ Item List cleared. Paste your new items into v2_ITEMS, then click '🔄 Refresh Data Now' in the sidebar.")
        else:
            st.error("❌ Confirmation text doesn't match. Nothing was deleted.")
    st.markdown('</div>', unsafe_allow_html=True)

    # Full Reset — Clear Everything
    st.markdown('<div class="card" style="border-color:#6A0F0F;">', unsafe_allow_html=True)
    st.markdown('<div class="section-title" style="color:#E08080;">Full Reset — Clear Everything</div>', unsafe_allow_html=True)
    st.write("Wipes **both** the Item List/Beginning Inventory **and** all Daily Transactions. Use this for a complete fresh start — then paste your new items and beginning stocks into `v2_ITEMS`, and click '🔄 Refresh Data Now' in the sidebar to load them in.")
    confirm3 = st.text_input("Type **FULL RESET** to confirm", key="confirm_full")
    if st.button("💥 Full Reset (Clear Everything)", key="btn_full_reset"):
        if confirm3.strip().upper() == "FULL RESET":
            for sheet_name, headers in [(ITEMS_SHEET, ITEMS_HEADERS), (LOG_SHEET, LOG_HEADERS)]:
                ws = ensure_sheet(ss, sheet_name, headers)
                ws.clear()
                ws.append_row(headers)
            invalidate_cache()
            st.success("✅ Full reset complete. Paste your new items/beginning inventory into v2_ITEMS, then click '🔄 Refresh Data Now' in the sidebar.")
        else:
            st.error("❌ Confirmation text doesn't match. Nothing was deleted.")
    st.markdown('</div>', unsafe_allow_html=True)